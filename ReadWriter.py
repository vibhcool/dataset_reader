import csv
import openpyxl
import os

root_path = os.path.dirname(os.path.abspath(__file__))

def file_to_dict(filename):
    ''' main function: write file to dictionary '''

    if filename[-4:] == '.csv':
        data = read_csv(filename)
    elif filename[-5:] == '.xlsx' or filename[-4:] == '.xls':
        data = read_xl(filename)
    elif filename[-4:] == '.txt'
        data = read_txt(filename)
    else:
        print('error')
        data = {}

    return data

    
def read_csv(filename):
    ''' read csv file to dict '''

    with open(root_path + filename) as csvfile:
        csvreader = csv.reader(csvfile, delimiter=' ', quotechar='|')

        data = {}
        data_row = []

        i = 0
        for row in csvreader:
            data[i] = {}
            j = 0
            data_row = row[0].split(',')
            for row_cell in data_row:
                a = row_cell 
                data[i][j] = a if a != None else 'na'
                j += 1
            i += 1

    data['filename'] = filename

    return data


def read_xl(filename):
    ''' read excel file '''

    wb = openpyxl.load_workbook(root_path + filename)
    anotherSheet = wb.active
    ws = wb.worksheets[0]

    data = {}

    for i in wb.worksheets:
        #all rows and columns
        row_count = i.max_row
        col_count = i.max_column
        print(row_count, col_count)
        for row in range(1,row_count):
            data[row-1] = {}
            for col in range(1,col_count):
                a = i.cell(column=col, row=row).value
                data[row-1][col-1] = a if a != None else 'na'

    data['filename'] = filename

    return data

def read_txt_by_line(filename)

    data = {}
    with open(root_path + filename, 'r'):
        i = 0
        while (line = f.readline()):
            data[i] = line
            i += 1
    return data            

def read_txt(filename)

    data = {}
    with open(root_path + filename, 'r'):
        data['data'] = f.read()
        data['filename'] = filename
    return data            
'''
def write_to_db(data, client_name, file_id, db_object):
    ''' write data extracted to database(sqlite) '''
    #ques_paper=quesFile.objects.get(ques_paper_id=filename, client=client_name)
    for i in data:
        if str(i) == 'filename':
            break
        db_object.objects.create(
            question_id=file_id,
            question=data[i].get(1),
            option1=data[i].get(2),
            option2=data[i].get(3),
            option3=data[i].get(4),
            option4=data[i].get(5),
            answer=data[i].get(6),
            questionType=data[i].get(7),
        )
'''
def write_to_out(data, client_name, file_id)
    ''' write data extracted to output data '''
    for i in data:
        if str(i) == 'filename':
            break
        print(data[file_id,
            data[i].get(1),
            data[i].get(2),
            data[i].get(3),
            data[i].get(4),
            data[i].get(5),
            data[i].get(6),
            data[i].get(7),
        )
        
if __name__ == '__main__'
    file_to_db('quesformat1.xlsx','yoyo')

